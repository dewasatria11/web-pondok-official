from http.server import BaseHTTPRequestHandler
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from lib._supabase import supabase_client


class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        """
        GET /api/export_pendaftar_xlsx
        Response: Excel file download (.xlsx)
        """
        try:
            # Get Supabase client with service role for full access
            supa = supabase_client(service_role=True)

            # Query pendaftar table with computed boolean fields
            # Using CASE to create boolean flags for file existence
            result = (
                supa.table("pendaftar")
                .select("""
                    gelombang,
                    nisn,
                    namalengkap,
                    tanggallahir,
                    tempatlahir,
                    namaayah,
                    namaibu,
                    telepon_orang_tua,
                    rencanatingkat,
                    rencanaprogram,
                    alamatjalan,
                    desa,
                    file_akta,
                    file_ijazah,
                    file_foto,
                    file_kk,
                    file_bpjs
                """)
                .execute()
            )
            
            if not result.data:
                self.send_response(404)
                self.send_header("Content-Type", "application/json")
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(
                    b'{"ok": false, "error": "Tidak ada data pendaftar"}'
                )
                return

            # Transform data to match expected format
            rows = []
            for item in result.data:
                # Build alamat_lengkap (alamatjalan + desa separated by comma)
                alamat_parts = []
                if item.get('alamatjalan'):
                    alamat_parts.append(item['alamatjalan'].strip())
                if item.get('desa'):
                    alamat_parts.append(item['desa'].strip())
                alamat_lengkap = ', '.join(filter(None, alamat_parts))
                
                # Check if files exist (any non-empty value counts as file exists)
                def has_file(field_value):
                    """Check if file exists: not null, not empty, not 'null' string"""
                    if not field_value:
                        return False
                    str_value = str(field_value).strip().lower()
                    # Exclude common "empty" values
                    if str_value in ['', 'null', 'none', 'undefined']:
                        return False
                    return True
                
                has_file_akta = has_file(item.get('file_akta'))
                has_file_ijazah = has_file(item.get('file_ijazah'))
                has_file_foto = has_file(item.get('file_foto'))
                has_file_kk = has_file(item.get('file_kk'))
                has_file_bpjs = has_file(item.get('file_bpjs'))
                
                rows.append({
                    'gelombang': item.get('gelombang', ''),
                    'nisn': item.get('nisn', ''),
                    'nama': item.get('namalengkap', ''),
                    'tanggal_lahir': item.get('tanggallahir', ''),
                    'tempat_lahir': item.get('tempatlahir', ''),
                    'nama_ayah': item.get('namaayah', ''),
                    'nama_ibu': item.get('namaibu', ''),
                    'nomor_orangtua': item.get('telepon_orang_tua', ''),
                    'rencana_tingkat': item.get('rencanatingkat', ''),
                    'rencana_program': item.get('rencanaprogram', ''),
                    'alamat_lengkap': alamat_lengkap,
                    'has_file_akta': has_file_akta,
                    'has_file_ijazah': has_file_ijazah,
                    'has_file_foto': has_file_foto,
                    'has_file_kk': has_file_kk,
                    'has_file_bpjs': has_file_bpjs,
                })
            
            # Sort by rencana_program (case-insensitive, A-Z), then by nama
            rows.sort(key=lambda x: (str(x.get('rencana_program', '')).lower(), str(x.get('nama', ''))))

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Pendaftar"

            # Define headers (EXACT ORDER)
            headers = [
                'gelombang',
                'nisn',
                'nama',
                'tanggal_lahir',
                'tempat_lahir',
                'nama_ayah',
                'nama_ibu',
                'nomor_orangtua',
                'rencana_tingkat',
                'rencana_program',
                'alamat_lengkap',
                'file_akta',  # FIXED: was 'file_akte', should be 'file_akta' to match database
                'file_ijazah',
                'file_foto',
                'file_kk',
                'file_bpjs'
            ]

            # Write headers
            ws.append(headers)

            # Style header row
            header_fill = PatternFill(start_color="0F9D58", end_color="0F9D58", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Write data rows
            for row_data in rows:
                row_values = []
                for header in headers:
                    # Map has_file_* to file_* with YA/TIDAK
                    if header.startswith('file_'):
                        boolean_key = 'has_' + header
                        value = 'YA' if row_data.get(boolean_key, False) else 'TIDAK'
                        
                    else:
                        value = row_data.get(header, '')
                    
                    # Handle None
                    if value is None:
                        value = ''
                    
                    # Handle date formatting
                    if header == 'tanggal_lahir' and value and value != '':
                        try:
                            # Parse date if string
                            if isinstance(value, str):
                                # Try YYYY-MM-DD format
                                date_obj = datetime.strptime(value[:10], '%Y-%m-%d')
                                value = date_obj
                        except:
                            pass
                    
                    row_values.append(value)
                
                ws.append(row_values)

            # Apply styling to all cells
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    cell.border = thin_border
                    
                    header_name = headers[col_idx - 1]
                    
                    # Text format for NISN and phone number (preserve leading zeros)
                    if header_name in ['nisn', 'nomor_orangtua']:
                        cell.number_format = '@'  # Text format
                    
                    # Date format
                    elif header_name == 'tanggal_lahir':
                        if isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                    
                    # Wrap text for long text columns
                    if header_name in ['alamat_lengkap', 'nama']:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Auto-fit columns
            for col_idx, col in enumerate(ws.columns, start=1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                
                for cell in col:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set column width with limits
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze pane at A2 (header row visible when scrolling)
            ws.freeze_panes = 'A2'

            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            excel_data = excel_buffer.read()

            # Generate filename
            today = datetime.now().strftime('%Y%m%d')
            filename = f"pendaftar_{today}.xlsx"

            # Send Excel response
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Content-Length', str(len(excel_data)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Cache-Control', 'no-cache')
            self.end_headers()
            self.wfile.write(excel_data)

            print(f"✓ Excel exported: {filename} ({len(rows)} rows)")

        except Exception as e:
            print(f"Error in export_pendaftar_xlsx: {e}")
            import traceback
            traceback.print_exc()
            
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(
                f'{{"ok": false, "error": "{str(e)}"}}'.encode('utf-8')
            )

    def do_OPTIONS(self):
        """Handle CORS preflight"""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

